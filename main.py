import pandas as pd
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import filedialog

def abrir_arquivo():
    global caminho
    global df
    
    root = tk.Tk()
    root.wm_attributes('-topmost', True)
    root.withdraw()

    caminho = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if caminho:
        df = pd.read_excel(caminho)
        print("Arquivo escolhido:", caminho)
    else:
        print("Nenhum arquivo escolhido.")
        main()
def ver_usuarios():
    if 'df' not in globals():
        print("Nenhum arquivo carregado, por favor abra um arquivo")
        abrir_arquivo()
    print(df)

def criar_usuario(nome, email, senha):
    if 'df' not in globals():
        print("Nenhum arquivo carregado, por favor abra um arquivo")
        abrir_arquivo()
    if df['Email'].str.contains(email).any():
        print("O email informado já existe!")
    else:
        novo_usuario={'Nome': nome.title(), 'Email': email, 'Senha': senha}
        df.loc[len(df)] = novo_usuario
        print ("usuario criado")

def atualizar_senha(email, nova_senha):
    if 'df' not in globals():
        print("Nenhum arquivo carregado, por favor abra um arquivo")
        abrir_arquivo()
    df.loc[df['Email'] == email, 'Senha'] = nova_senha
    print ("Senha atualizada")
def deletar_usuario(email):
    if 'df' not in globals():
        print("Nenhum arquivo carregado, por favor abra um arquivo")
        abrir_arquivo()
    if df['Email'].str.contains(email).any():
        index = df[df['Email'] == email].index
        df.drop(index, inplace=True)
        print ("Usuario deletado")
    else:
        print("Usuario nao encontrado!")

def buscar_usuario(email):
    if 'df' not in globals():
        print("Nenhum arquivo carregado, por favor abra um arquivo")
        abrir_arquivo()
    if df['Email'].str.contains(email).any():
        print(df.loc[df['Email'] == email])
    else:
        print("Usuario nao encontrado!")

def salvar_planilha():
    if 'df' not in globals():
        print("Nenhum arquivo carregado, por favor abra um arquivo")
        abrir_arquivo()
    from openpyxl import load_workbook
    df.to_excel(caminho, index=False)
    wb = load_workbook(caminho)
    ws = wb.active

    #formatador automatico de excel
    for colunas_celulas in ws.columns:
        largura = max(len(str(celula.value)) for celula in colunas_celulas)
        ws.column_dimensions[get_column_letter(colunas_celulas[0].column)].width = largura + 2
        for celula in colunas_celulas:
            celula.alignment = Alignment(wrap_text=True)
    wb.save(caminho)
    print("Planilha salva com sucesso!")


def main():
    while True:
        print("\nMenu:")
        print("1 - Escolher arquivo")
        print("2 - Ver usuarios")
        print("3 - Criar usuario")
        print("4 - Atualizar senha")
        print("5 - Deletar usuario")
        print("6 - Buscar usuario")
        print("7 - Salvar planilha, use com o excel fechado!")
        print("0 - Sair")
        try:
            opcao = int(input("Digite a opcao: "))
        except ValueError:
            print("Opção inválida! Por favor digite um numero.")
            continue
        if opcao == 0:
            break
        if opcao == 1:
            abrir_arquivo()
        elif opcao == 2:
            ver_usuarios()
        elif opcao == 3:
            nome = input("Digite o nome: ")
            email = input("Digite o email: ")
            senha = input("Digite a senha: ")
            criar_usuario(nome, email, senha)
        elif opcao == 4:
            email = input("Digite o email: ")
            nova_senha = input("Digite a nova senha: ")
            atualizar_senha(email, nova_senha)
        elif opcao == 5:
            email = input("Digite o email: ")
            deletar_usuario(email)
        elif opcao == 6:
            email = input("Digite o email: ")
            buscar_usuario(email)
        elif opcao == 7:
            salvar_planilha()
        else:
            print("Valor invalida! Por favor digite um numero de 0 a 6")


if __name__ == "__main__":
    main()

